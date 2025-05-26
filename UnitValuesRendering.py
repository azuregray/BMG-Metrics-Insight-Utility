'''
------------------------------------------------------------- DOCUMENTATION STARTS HERE

[SCRIPT NAME]   UnitValuesRendering.py

[SCRIPT TYPE/DEPENDENCIES]  No External Script Dependencies; Excel Template File - "Results-ExportTemplate.xlsx"

[DESCRIPTION]   This python script combines the functionalities of both Input & Output Files Processing >
                Targeting purely dimension layers of each DXF file >
                Parsing Dimensional Attributes values from modelspace entities >
                Post Processing parsed values > Manage workspace cleanUp >
                Present the results in a seamless fashion > Export to a structured Excel File on Demand.

[LIBRARIES USED]    (RUNNING WITH Web Interface)    os, re, tempfile, shutil, time.sleep, ezdxf, openpyxl.load_workbook
                    (RUNNING WITH CLI)              + tkinter.filedialog, ctypes.windll

[FUNCTIONS]     gracefulErrors(errorMessage),
                renderValues(dxfPath, dxfType),
                pathCorresponder(inputDirPath, outputDirPath),
                harvesterFunc(inputDirPath, outputDirPath),   --> MAIN FUNCTION
                excelProcessor(exportableData)

[NOTE]  1.  This code runs only on DXF files.
        2.  Please make sure, filenames are exactly identical in both the folders before running this script.
        3.  Please make sure you have the Excel Template file named "Results-ExportTemplate.xlsx" in the same directory as the script.
        4.  Main Section is also included for better UI/UX, while using with CLI. (Refer section starting Line:213)

------------------------------------------------------------- DOCUMENTATION ENDS HERE
'''

import os, tempfile

def gracefulErrors(errorMessage, exitRequired=False):
    if exitRequired:
        if __name__ == '__main__':
            from time import sleep
            os.system('cls')
            print("\n\n:::::::: ❌ The system faced an error. ::::::::")
            sleep(0.8)
            print(f'[Error Message] ::: ⚠️ {errorMessage}')
            sleep(1)
            print("\n\n:::::::: Abort All Operations & Exit ::::::::\n")
            sleep(0.8)
            for letter in 'Press [ENTER] to exit...':
                    print(letter, flush=True, end='')
                    sleep(0.05)
            input()
            print()
            exit(1)
        else:
            raise Exception(errorMessage)
    elif not exitRequired:
        print(f'\n\n❌ WARNING ::: {errorMessage}\n\n')
        if (('errorLogs' in locals()) or ('errorLogs' in globals())) and (__name__ == '__main__'):
            errorLogs.append(f'❌ WARNING ::: {errorMessage}')

def renderValues(dxfPath, dxfType):
    import ezdxf
    
    preReturnableList = []
    if dxfType.lower() == 'input':
        import re

        filters = [
            r'\{\\H1\.88x;\(}',
            r'\{\\H1\.88x;\)}',
        ]
        
        try:
            doc = ezdxf.readfile(dxfPath)
            msp = doc.modelspace()
            for mtext_entity in msp.query('MTEXT[layer=="41"]'):
                try:
                    value = getattr(mtext_entity, 'text')
                    if isinstance(value, (int, float, str)):
                        cleaned_data = re.sub('|'.join(filters), '', value)
                        cleaned_data = re.sub(r'[^;]+;', '', cleaned_data)
                        cleaned_data = cleaned_data.replace(',', '.')
                        if cleaned_data and not cleaned_data.startswith('.'):
                            cleaned_data = cleaned_data.strip()
                            cleaned_data = re.sub(r'%%[pdc]', '', cleaned_data)
                            cleaned_data = cleaned_data.strip()
                            if re.match(r'^-?\d*\.?\d+$', cleaned_data):
                                preReturnableList.append(f'{float(cleaned_data):.1f}')
                except:
                    continue
            if preReturnableList:
                returnableList = sorted(set(float(value) for value in preReturnableList if (value != '0.0')))
        except Exception as e:
            gracefulErrors(f"Error processing Input DXF File >> {os.path.basename(dxfPath)} :: {e}")
            return []
    elif dxfType.lower() == 'output':
        try:
            doc = ezdxf.readfile(dxfPath)
            msp = doc.modelspace()
            for entity in msp.query('DIMENSION'):
                for attr_name in dir(entity.dxf):
                    if attr_name == 'actual_measurement':
                        try:
                            value = entity.dxf.get(attr_name)
                            preReturnableList.append(f'{float(str(value).strip()):.1f}')
                        except:
                            continue
            if preReturnableList:
                returnableList = sorted(set(float(value.strip()) for value in preReturnableList if (value != '1.0')))
        except Exception as e:
            gracefulErrors(f"Error processing Output DXF File >> {os.path.basename(dxfPath)} :: {e}")
            return []
    return returnableList

def pathCorresponder(inputDirPath, outputDirPath):
    inputFilesList = sorted([file for file in os.listdir(inputDirPath) if file.lower().endswith('.dxf')])
    outputFilesList = sorted([file for file in os.listdir(outputDirPath) if file.lower().endswith('.dxf')])
    
    if (len(inputFilesList) * len(outputFilesList)) == 0:
        gracefulErrors('Trouble finding suitable files. Please make sure there are mutually named files in both directories.', exitRequired=True)
    
    mutualFilesList = [file for file in inputFilesList if file in outputFilesList]
    nonMutualFilesList = [file for file in inputFilesList if file not in outputFilesList] + [file for file in outputFilesList if file not in inputFilesList]
    
    returnableList = []
    
    for index in range(len(mutualFilesList)):
        filename = str(mutualFilesList[index])
        inputFileName = f'{inputDirPath}/{filename}'
        outputFileName = f'{outputDirPath}/{filename}'
        flag = 'matched'
        if not (os.path.exists(inputFileName) and os.path.exists(outputFileName)):
            gracefulErrors(f'There was trouble finding corresponding file {filename} in both the selected directories.\nMake sure no program is working with the directories and wait until this program is done.')
            continue
        else:
            returnableList.append([filename, inputFileName, outputFileName, flag])
    
    for file in (nonMutualFilesList):
        returnableList.append([str(file),'unmatched'])
    
    return returnableList

def harvesterFunc(inputDirPath, outputDirPath, progress_callback=None):
    finalResults = []
    corresponderList = pathCorresponder(inputDirPath, outputDirPath)
    
    if progress_callback:
        progressTotal = len(corresponderList)
        progressCounter = 0
    
    for fileListItem in corresponderList:
        if len(fileListItem) == 4 and fileListItem[3].lower() == 'matched':
            fileName = fileListItem[0]
            inputValuesList = renderValues(fileListItem[1], dxfType='Input')
            outputValuesList = renderValues(fileListItem[2], dxfType='Output')
            if (len(inputValuesList) * len(outputValuesList)) == 0:
                similarityVerdict = 'False'
            else:
                similarityVerdict = 'True' if all([value in inputValuesList for value in outputValuesList]) else 'False'
            inputValuesList, outputValuesList = sorted(inputValuesList, reverse=True), sorted(outputValuesList, reverse=True)
            outputValuesList = [val if val in outputValuesList else '' for val in inputValuesList] + [value for value in outputValuesList if value not in inputValuesList]
            finalResults.append([fileName[:-4], inputValuesList, outputValuesList, similarityVerdict])
            if progress_callback:
                progressCounter += 1
                progress_callback(progressCounter/progressTotal)
        elif len(fileListItem) == 2 and fileListItem[1].lower() == 'unmatched':
            fileName = fileListItem[0]
            similarityVerdict = 'unmatched'
            finalResults.append([fileName[:-4], [], [], similarityVerdict])
            if progress_callback:
                progressCounter += 1
                progress_callback(progressCounter/progressTotal)
    
    return finalResults

def excelProcessor(exportableData, exportTemplatePath=os.path.join(os.path.dirname(__file__), 'Results-ExportTemplate.xlsx'), exportDir=tempfile.gettempdir()):
    from openpyxl import load_workbook
    
    try:
        wb = load_workbook(exportTemplatePath)
    except FileNotFoundError:
        gracefulErrors(f'Excel template file not found at >> {exportTemplatePath}', exitRequired=True)
    except Exception as e:
        gracefulErrors(f'Error loading Excel template :: {e}', exitRequired=True)
    
    ws = wb["MainSheet"]
    start_row = 2

    for i, (material_id, input_vals, output_vals, verdict) in enumerate(exportableData):
        input_row = start_row + i * 2
        output_row = input_row + 1

        ws.cell(row=input_row, column=1, value=material_id)
        if (verdict.lower() == 'unmatched'):
            ws.cell(row=input_row, column=2, value=f'[{verdict.upper()}]')
        else:
            ws.cell(row=input_row, column=2, value=verdict.upper())

        if ((len(input_vals) * len(output_vals)) * (verdict.lower() != 'unmatched')):
            for j, val in enumerate(input_vals):
                ws.cell(row=input_row, column=3 + j, value=float(val))
            
            for j, val in enumerate(output_vals):
                if val == '':
                    continue
                ws.cell(row=output_row, column=3 + j, value=float(val))

    savePath = os.path.join(exportDir, "Results-EXPORT.xlsx")
    
    try:
        wb.save(savePath)
    except Exception as e:
        gracefulErrors(f'Error saving Excel file to >> {savePath} :: {e}', exitRequired=True)
    
    return savePath

if __name__ == '__main__':
    import tkinter.filedialog as fd
    from time import sleep
    from ctypes import windll
    
    windll.shcore.SetProcessDpiAwareness(1)
    errorLogs = []
    
    os.system('cls')
    print(":::::::: Let's Start ::::::::")
    sleep(1)
    
    os.system('cls')
    print(":::::::: Select a Input Files Folder to use.. in the prompt that appears now. ::::::::")
    sleep(0.8)
    inputFilesDir = fd.askdirectory(title='Select the folder with Input DXF Files.')
    os.system('cls')
    
    os.system('cls')
    print(":::::::: Select a Output Files Folder to use.. in the prompt that appears now. ::::::::")
    sleep(0.8)
    outputFilesDir = fd.askdirectory(title='Select the folder with Output DXF Files.')
    os.system('cls')
    
    print(":::::::: Scanning your folders for DXF Files... ::::::::")
    finalResults = harvesterFunc(inputFilesDir, outputFilesDir)
    
    print('\n\n:::: FINAL RESULTS ::::')
    for index, (fileName, inputValuesList, outputValuesList, similarityVerdict) in enumerate(finalResults, start=1):
            print(f'\n:::: File #{index} ::::')
            print(f'Processed File Name :::: {fileName}')
            if similarityVerdict.lower() != 'unmatched':
                print(f'Input Values :::: {inputValuesList}')
                print(f'Output Values :::: {outputValuesList}')
                if similarityVerdict == 'True':
                    print('Are they same? :::: ✅✅ YES ✅✅\n')
                elif similarityVerdict == 'False':
                    print('Are they same? :::: ❌❌ NO ❌❌\n')
            elif similarityVerdict.lower() == 'unmatched':
                print('Condition :::: ⚠️⚠️ Matching File Not Found ⚠️⚠️\n')
    
    print('\n\nData has been processed from your files.')
    
    if errorLogs:
        print('\n\n\n:::::::: But, There were also some non-critical errors. ::::::::\n')
        for index, logMessage in enumerate(errorLogs, start=1):
            print(f'{index}. {logMessage}')
    
    for letter in '\nPress [ENTER] to exit.\nOr You can also type [YES] to export the results into an Excel file.\n':
        print(letter, flush=True, end='')
        sleep(0.04)
    
    keypress = input()
    if keypress.lower() == 'yes':
        print(":::::::: Choose a folder to export the Excel in the popup that appears now. ::::::::")
        exportDir = fd.askdirectory(title='Select the folder saving your Excel Export File.')
        exportTemplatePath = os.path.join(os.path.dirname(__file__), 'Results-ExportTemplate.xlsx')
        try:
            exportedPath = excelProcessor(
                exportableData=finalResults,
                exportTemplatePath=exportTemplatePath,
                exportDir=exportDir
            )
        except Exception as e:
            gracefulErrors(f'There was an issue while exporting your results at excelProcessor() :: {e}', exitRequired=True)
        print(f'\n\nYour Excel file has been exported to --> {exportedPath}.')
        print('\n\n:::::::: Now Opening your Excel Export. ::::::::')
        sleep(1)
        os.startfile(exportedPath)
    else:
        print('Thanks for using the script. Exiting Now.')
    
    print()

